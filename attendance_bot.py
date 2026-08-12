import base64
import calendar
from datetime import datetime
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import json
import os
import smtplib
import sys
import time

from dotenv import find_dotenv, load_dotenv
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.sync_api import Playwright, sync_playwright

# Force load .env file dynamically
load_dotenv(find_dotenv(), override=True)

USERNAME = os.getenv("HR_USERNAME") or os.getenv("USERNAME") or ""
PASSWORD = os.getenv("HR_PASSWORD") or os.getenv("PASSWORD") or ""

# Email Settings
SENDER_EMAIL = os.getenv("SENDER_EMAIL") or ""
SENDER_APP_PASSWORD = os.getenv("SENDER_APP_PASSWORD") or ""
RECEIVER_EMAIL = os.getenv("RECEIVER_EMAIL") or ""


def extract_attendance_from_image(image_path: str) -> list[dict]:
    """Uses OpenAI Vision to extract daily records and color-coded status."""
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise ValueError("❌ OPENAI_API_KEY missing from environment!")

    client = OpenAI(api_key=api_key)

    with open(image_path, "rb") as image_file:
        base64_image = base64.b64encode(image_file.read()).decode("utf-8")

    prompt = (
        "Analyze all daily entries visible in this attendance calendar screenshot.\n"
        "Pay special attention to the color coding and tags of each date block:\n"
        "- GREEN color indicates 'Present'.\n"
        "- RED color indicates 'Absent'.\n"
        "- GREY / OFF-COLOR or tagged as 'WO' / 'PH' indicates 'Week Off' or 'Public Holiday'.\n"
        "- ANY OTHER COLOR (or multiple event highlights/partial punches) indicates 'Multiple Events'.\n\n"
        "Return a JSON object with a key 'records' which is an array of objects.\n"
        "Each object must have these exact keys:\n"
        "- 'Date' (e.g., 'Aug 1')\n"
        "- 'Status' ('Present', 'Absent', 'Week Off', or 'Multiple Events')\n"
        "- 'In Time'\n"
        "- 'Out Time'\n"
        "- 'Working Hours'"
    )

    response = client.chat.completions.create(
        model="gpt-4o",
        response_format={"type": "json_object"},
        messages=[
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {
                        "type": "image_url": {
                            "url": f"data:image/png;base64,{base64_image}"
                        },
                    },
                ],
            }
        ],
        temperature=0.1,
    )

    result_text = response.choices[0].message.content
    data = json.loads(result_text)

    if isinstance(data, dict):
        for key, val in data.items():
            if isinstance(val, list):
                return val
    return data if isinstance(data, list) else []


def generate_custom_email_summary(records: list[dict]) -> str:
    """Generates the specific monthly breakdown requested."""
    api_key = os.getenv("OPENAI_API_KEY")
    client = OpenAI(api_key=api_key)

    today = datetime.now()
    days_in_month = calendar.monthrange(today.year, today.month)[1]
    month_name = today.strftime("%B %Y")

    prompt = (
        f"Month: {month_name}\n"
        f"Total days in this month: {days_in_month}\n"
        f"Parsed calendar records: {records}\n\n"
        "Generate a structured summary following these exact metrics:\n"
        "1. Total Number of Days in this Month\n"
        "2. Number of Present Days\n"
        "3. Number of Absent Days\n"
        "4. Number of Week Offs / Holidays\n"
        "5. Number of Multiple Events Days (Irregular Attendance)\n"
        "6. Calculate Total Working Hours Logged\n\n"
        "7. REGULARIZATION & LEAVE ACTION SECTION:\n"
        "   - List all ABSENT DATES separately.\n"
        "   - List all MULTIPLE EVENT DATES separately.\n"
        "   - Calculate Total Count of Days Needing Regularization/Leave (Absent + Multiple Events).\n\n"
        "Format the summary cleanly for email reading like this:\n"
        "📊 MONTHLY ATTENDANCE SUMMARY REPORT\n"
        "=========================================\n"
        "📅 Month: [Month Year]\n"
        "📆 Total Days in Month: [Total Days]\n\n"
        "📈 BREAKDOWN:\n"
        "-----------------------------------------\n"
        "✅ Present Days: X days\n"
        "❌ Absent Days: X days\n"
        "🏖️ Week Offs / Holidays: X days\n"
        "⚠️ Multiple Events / Irregular Days: X days\n"
        "⏱️ Total Working Hours Logged: XX hrs XX mins\n\n"
        "🚨 REGULARIZATION & LEAVE ACTION REQUIRED:\n"
        "-----------------------------------------\n"
        "🔴 Absent Dates (Apply Leave / Regularization):\n"
        "   - [Dates or 'None']\n\n"
        "🟠 Multiple Event Dates (Regularization Needed):\n"
        "   - [Dates or 'None']\n\n"
        "📌 TOTAL DAYS NEEDING ACTION: X days"
    )

    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[{"role": "user", "content": prompt}],
        temperature=0.2,
    )

    return response.choices[0].message.content


def send_email_report(summary: str, excel_path: str):
    """Sends the summary via email and attaches the Excel report."""
    if not SENDER_EMAIL or not SENDER_APP_PASSWORD or not RECEIVER_EMAIL:
        print("⚠️ Email credentials missing in .env! Skipping email delivery.")
        return

    today_str = datetime.now().strftime("%Y-%m-%d")

    msg = MIMEMultipart()
    msg["From"] = SENDER_EMAIL
    msg["To"] = RECEIVER_EMAIL
    msg["Subject"] = f"📊 Monthly Attendance & Regularization Report - {today_str}"

    # Body
    msg.attach(MIMEText(summary, "plain"))

    # Attachment
    if os.path.exists(excel_path):
        with open(excel_path, "rb") as f:
            part = MIMEApplication(f.read(), Name=os.path.basename(excel_path))
            part["Content-Disposition"] = (
                f'attachment; filename="{os.path.basename(excel_path)}"'
            )
            msg.attach(part)

    try:
        print("📧 Connecting to SMTP server...")
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(SENDER_EMAIL, SENDER_APP_PASSWORD)
            server.send_message(msg)
        print(f"✅ Email report successfully sent to {RECEIVER_EMAIL}!")
    except Exception as e:
        print(f"❌ Failed to send email report: {e}")


def run(playwright: Playwright):
    if not USERNAME or not PASSWORD:
        print("❌ Error: Missing credentials in environment variables!")
        sys.exit(1)

    os.makedirs("screenshots", exist_ok=True)

    # Launch browser configured for server environments
    browser = playwright.chromium.launch(headless=True)
    context = browser.new_context(viewport={"width": 1920, "height": 1080})
    page = context.new_page()

    print("🔑 Navigating to login page...")
    page.goto("https://gleneagles.myadrenalin.com/AdrenalinMax/#/", timeout=60000)

    # Fill login details safely
    page.get_by_role("textbox", name="User ID").fill(str(USERNAME))
    page.get_by_role("textbox", name="Password").fill(str(PASSWORD))
    page.get_by_role("button", name="Login").click()

    # Wait for post-login dashboard loading
    print("⏳ Waiting for dashboard load...")
    page.wait_for_timeout(7000)
    page.screenshot(path="screenshots/debug_after_login.png")

    # Resilient Navigation using force clicks and broad text matching
    print("📍 Clicking Attendance sidebar menu...")
    menu_selectors = [
        "text=Attendance, claims & requests",
        "text=Attendance & leave",
        "text=Attendance",
        "[aria-label*='Attendance']",
    ]

    clicked = False
    for selector in menu_selectors:
        try:
            loc = page.locator(selector).first
            if loc.is_visible(timeout=3000):
                loc.click(force=True, timeout=5000)
                clicked = True
                print(f"✅ Clicked menu via selector: {selector}")
                break
        except Exception:
            continue

    if not clicked:
        # Final fallback: click any element containing 'Attendance'
        page.locator("*:has-text('Attendance')").last.click(force=True)

    page.wait_for_timeout(3000)

    # Sub-menu click
    try:
        page.locator("text=Attendance & leave").first.click(force=True, timeout=5000)
    except Exception:
        pass

    # Handle iframe content
    print("🖼️ Accessing Attendance frame...")
    frame_element = page.wait_for_selector("#commonFormRender", timeout=20000)
    frame = frame_element.content_frame()

    # Wait for Show Time checkbox
    frame.get_by_role("checkbox", name="Show time").wait_for(timeout=15000)
    frame.get_by_role("checkbox", name="Show time").check()

    # Click Refresh
    frame.get_by_text("Refresh").click()

    # Wait for calendar content to render
    page.wait_for_timeout(5000)

    # Save screenshot directly from iframe
    today = datetime.now().strftime("%Y-%m-%d")
    screenshot_path = f"screenshots/attendance_{today}.png"

    frame.locator("body").screenshot(path=screenshot_path)
    print(f"✅ Calendar screenshot saved: {screenshot_path}")

    # Process screenshot using AI Vision
    print("🤖 Processing calendar colors and attendance with OpenAI Vision...")
    records = extract_attendance_from_image(screenshot_path)

    # Generate Actionable Summary
    print("📊 Generating Detailed Attendance Summary Report...")
    summary = generate_custom_email_summary(records)

    print("\n--- GENERATED EMAIL SUMMARY ---")
    print(summary)
    print("--------------------------------\n")

    # Create Excel workbook
    wb = Workbook()

    # Sheet 1: Logs
    ws_data = wb.active
    ws_data.title = "Attendance Logs"
    ws_data.append(["Date", "Status", "In Time", "Out Time", "Working Hours"])

    for item in records:
        ws_data.append(
            [
                item.get("Date", ""),
                item.get("Status", ""),
                item.get("In Time", ""),
                item.get("Out Time", ""),
                item.get("Working Hours", ""),
            ]
        )

    # Header styling
    header_fill = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    for cell in ws_data[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws_data.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_data.column_dimensions[col_letter].width = max(max_len + 4, 15)

    # Sheet 2: Action Summary Report
    ws_summary = wb.create_sheet(title="Action Report")
    ws_summary.append(["Attendance Action & Regularization Report"])
    ws_summary.cell(row=1, column=1).font = Font(
        name="Calibri", size=14, bold=True, color="1F4E79"
    )
    ws_summary.append([])

    for line in summary.split("\n"):
        if line.strip():
            ws_summary.append([line.strip()])

    ws_summary.column_dimensions["A"].width = 80

    # Save Excel
    os.makedirs("reports", exist_ok=True)
    excel_path = os.path.join("reports", f"Attendance_{today}.xlsx")
    wb.save(excel_path)
    print(f"📁 Excel Report generated at: {excel_path}")

    # Send Email Report
    send_email_report(summary, excel_path)

    context.close()
    browser.close()


if __name__ == "__main__":
    with sync_playwright() as playwright:
        run(playwright)